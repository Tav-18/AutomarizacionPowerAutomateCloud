from __future__ import annotations

import os
import re
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# =========================
# 1) Headers
# =========================
DETAIL_HEADERS = [
    "Title",
    "Internal Path",
    "Target",
    "Impact area",
    "Suggestion",
]

AZURE_HEADERS = [
    "ID",
    "State",
    "Tags",
    "Work Item Type",
    "Title",
    "Internal Path",
    "Target",
    "Impact area",
    "Suggestion",
    "Incidents Repeats",
    "Review type",
    "Developer",
    "Iteration Path",
]


# =========================
# 2) Helpers: Flow / Action formatting
# =========================
def flow_base(flow_name: str) -> str:
    """
    'AutonomousAgentsIsaGPO_2-79C3BC7D-...json' -> 'AutonomousAgentsIsaGPO_2'
    """
    if not flow_name:
        return ""
    base = os.path.basename(flow_name).strip()
    base = re.sub(r"\.json$", "", base, flags=re.IGNORECASE)
    if "-" in base:
        base = base.split("-", 1)[0]
    return base.strip()


def action_pretty(action_name: str) -> str:
    """
    Reglas solicitadas:
    - '_' -> ' '
    - quitar prefijos tipo "Get data on RPA ..." (y variantes)
    - opcional: quitar "content" al final
    """
    if not action_name:
        return ""

    s = action_name.replace("_", " ").strip()

    prefixes = [
        "Get data on RPA ",
        "Get data on rpa ",
        "Get data on ",
        "Get data ",
    ]
    s_low = s.lower()
    for p in prefixes:
        if s_low.startswith(p.lower()):
            s = s[len(p):].strip()
            break

    s = re.sub(r"\bcontent\b$", "", s, flags=re.IGNORECASE).strip()
    return s


def make_target(flow_name: str, action_name: str) -> str:
    """
    Formato:
    'AutonomousAgentsIsaGPO_2 / Información de empleo API'
    """
    fb = flow_base(flow_name)
    ap = action_pretty(action_name)
    return f"{fb} / {ap}".strip(" /")

def make_action_target(action_name: str) -> str:
    """
    Solo el nombre amigable de la actividad.
    Ejemplo:
    'Initialize_variable_-__ObjM004' -> 'Initialize variable - ObjM004'
    """
    return action_pretty(action_name)


def build_detail_internal_path(f: dict) -> str:
    flow_name = (f.get("flow_name") or "").strip()
    action_name = (f.get("action_name") or "").strip()

    fb = flow_base(flow_name)
    ap = action_pretty(action_name)

    return " / ".join([x for x in [fb, ap] if x])


def make_action_target(action_name: str) -> str:
    return action_pretty(action_name)


def build_findings_rows(findings: List[dict]) -> List[List[str]]:
    rows: List[List[str]] = []

    for f in findings:
        rule_name = (f.get("rule_name") or "").strip()
        action_name = (f.get("action_name") or "").strip()

        mapping = map_rule(rule_name)

        rows.append([
            mapping["title"],
            build_detail_internal_path(f),
            make_action_target(action_name),
            mapping["impact_area"],
            mapping["suggestion"],
        ])

    return rows

# =========================
# 3) Mapping: rule_name -> catálogo
# =========================
RULE_CATALOG: Dict[str, Dict[str, object]] = {
    "Hardcode": {
        "impact_area": "C - Hard Code",
        "title": "Hard Code",
        "suggestion_es": (
            "Evitar valores sensibles hardcodeados. "
            "Mover a un repositorio seguro (por ejemplo variables de entorno protegidas, "
            "Key Vault o connection references) según el estándar."
        ),
    },
    "Parametrizable": {
        "impact_area": "M - Parametrizable",
        "title": "Parametrizable",
        "suggestion_es": (
            "Mover este valor a un mecanismo parametrizable (por ejemplo variables de entorno, "
            "parámetros o configuración central) según el estándar."
        ),
    },
    "Nomenclatura de actividades": {
        "impact_area": "L - Nombre de Actividades",
        "title": "Nombre de Actividades",
        "suggestion_es": (
            "Renombrar la actividad a un título descriptivo y único, alineado al manual."
        ),
    },
    "Manejo de errores (RunAfter)": {
        "impact_area": "E - Manejo de errores",
        "title": "Manejo de errores",
        "suggestion_es": (
            "Configurar RunAfter para estados failed, timedOut o canceled y agregar manejo controlado."
        ),
    },
    "Retrasos (Delay y Wait)": {
        "impact_area": "E - Retrasos",
        "title": "Retrasos",
        "suggestion_es": (
            "Evitar demoras fijas; preferir condiciones o esperas robustas, o parametrizar la demora."
        ),
    },
    "Nomenclatura de variables": {
        "impact_area": "L - Nombre de Variables",
        "title": "Nombre de Variables",
        "suggestion_es": (
            "Renombrar la variable siguiendo el estándar Cloud: "
            "Bln/Int/Flt/Str/Obj/Arr + UpperCamelCase."
        ),
    },
    "Prefijos variables / parámetros": {
        "impact_area": "L - Nombre de Argumentos",
        "title": "Nombre de Argumentos",
        "suggestion_es": (
            "Renombrar parámetros usando in_, out_ o io_ y el tipo de dato según el estándar definido."
        ),
    },
    "Nomenclatura de flujos": {
        "impact_area": "L - Nombre de proyecto",
        "title": "Nombre de proyecto",
        "suggestion_es": (
            "Renombrar el flujo usando una convención clara, uniforme y alineada al manual."
        ),
    },
    "Condición IF": {
        "impact_area": "M - Condición IF",
        "title": "Condición IF",
        "suggestion_es": (
            "Revisar la estructura de la condición para que la rama verdadera contenga la lógica principal y sea legible."
        ),
    },
    "Comentarios descriptivos": {
        "impact_area": "M - Descripciones",
        "title": "Descripciones",
        "suggestion_es": (
            "Agregar notas o descripciones suficientes en acciones y ámbitos para facilitar mantenimiento."
        ),
    },
}


def map_rule(rule_name: str) -> Dict[str, str]:
    if not rule_name:
        rule_name = "Finding"

    mapping = RULE_CATALOG.get(rule_name)
    if mapping:
        return {
            "impact_area": str(mapping["impact_area"]),
            "title": str(mapping["title"]),
            "suggestion": str(mapping["suggestion_es"]),
        }

    return {
        "impact_area": "M - Requerimientos",
        "title": rule_name,
        "suggestion": "Revisar la incidencia y alinearla al manual de buenas prácticas interno.",
    }


# =========================
# 4) Internal Path
# =========================
def build_internal_path(f: dict) -> str:
    """
    Preferencia:
    1) flow_file_relpath :: json_path
    2) flow_file_relpath
    3) json_path
    """
    json_path = (f.get("json_path") or "").strip()
    flow_rel = (f.get("flow_file_relpath") or "").strip()

    json_path_clean = json_path.replace("_", " ")

    if flow_rel and json_path_clean:
        return f"{flow_rel} :: {json_path_clean}"
    if flow_rel:
        return flow_rel
    return json_path_clean

def build_azure_internal_path(f: dict) -> str:
    """
    Para la hoja azure_ready:
    solo el nombre amigable de la actividad,
    sin json_path técnico tipo actions....
    """
    action_name = (f.get("action_name") or "").strip()
    return action_pretty(action_name)



# =========================
# 5) Build rows
# =========================
def build_findings_rows(findings: List[dict]) -> List[List[str]]:
    """
    Hoja details:
    - una fila por incidencia
    - Target = solo actividad
    - Internal Path = Flow / Actividad :: json_path
    """
    rows: List[List[str]] = []

    for f in findings:
        rule_name = (f.get("rule_name") or "").strip()
        action_name = (f.get("action_name") or "").strip()

        mapping = map_rule(rule_name)

        title = mapping["title"]
        internal_path = build_detail_internal_path(f)
        target = make_action_target(action_name)
        impact_area = mapping["impact_area"]
        suggestion = mapping["suggestion"]

        rows.append([title, internal_path, target, impact_area, suggestion])

    return rows


def build_azure_internal_path(f: dict) -> str:
    """
    Para la hoja azure_ready:
    solo el nombre amigable de la actividad.
    """
    action_name = (f.get("action_name") or "").strip()
    return action_pretty(action_name)


def build_azure_like_rows(findings: List[dict]) -> List[List[str]]:
    grouped: Dict[tuple, Dict[str, object]] = {}

    for f in findings:
        rule_name = (f.get("rule_name") or "").strip()
        flow_name = (f.get("flow_name") or "").strip()
        action_name = (f.get("action_name") or "").strip()

        mapping = map_rule(rule_name)

        title = mapping["title"]
        internal_path = build_azure_internal_path(f)
        target = make_target(flow_name, action_name)
        impact_area = mapping["impact_area"]
        suggestion = mapping["suggestion"]

        key = (title, impact_area, suggestion)

        if key not in grouped:
            grouped[key] = {
                "title": title,
                "impact_area": impact_area,
                "suggestion": suggestion,
                "internal_paths": [],
                "targets": [],
                "count": 0,
            }

        item = grouped[key]

        if internal_path and internal_path not in item["internal_paths"]:
            item["internal_paths"].append(internal_path)

        if target and target not in item["targets"]:
            item["targets"].append(target)

        item["count"] += 1

    rows: List[List[str]] = []

    for _, item in grouped.items():
        rows.append([
            "",  # ID
            "Active",  # State
            "",  # Tags
            "Code Review",  # Work Item Type
            item["title"],
            "\n".join(item["internal_paths"]),
            "\n".join(item["targets"]),
            item["impact_area"],
            item["suggestion"],
            item["count"],
            "",  # Review type
            "",  # Developer
            "",  # Iteration Path
        ])

    rows.sort(key=lambda r: (r[4], r[7]))
    return rows

# =========================
# 6) Sheet writer
# =========================
def _write_sheet(ws, headers, rows, col_widths=None, wrap_cols=None):
    thin = Side(style="thin", color="D9DEE8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(color="FFFFFF", bold=True)
    body_font = Font(color="1F2937")

    ws.append(headers)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in rows:
        ws.append(row)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    if col_widths:
        for col_idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    if wrap_cols:
        for row in ws.iter_rows(min_row=2):
            for col_idx in wrap_cols:
                row[col_idx - 1].alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(rows) + 1)}"


# =========================
# 7) Export
# =========================
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def _write_sheet(ws, headers, rows, col_widths=None, wrap_cols=None):
    thin = Side(style="thin", color="D9DEE8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)

    ws.append(headers)

    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    for r in rows:
        ws.append(r)

    for row in ws.iter_rows(min_row=2, max_row=len(rows) + 1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

    if col_widths:
        for col_idx, w in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(rows) + 1)}"


def export_findings_to_xlsx(
    out_path,
    findings: List[dict],
    project_id: str = "",
) -> None:
    wb = Workbook()

    default_ws = wb.active
    wb.remove(default_ws)

    ws_details = wb.create_sheet("details")
    ws_azure = wb.create_sheet("azure_ready")

    detail_rows = build_findings_rows(findings)
    azure_rows = build_azure_like_rows(findings)

    detail_widths = {
        1: 28,
        2: 70,
        3: 46,
        4: 24,
        5: 90,
    }

    azure_widths = {
        1: 10,   # ID
        2: 14,   # State
        3: 18,   # Tags
        4: 18,   # Work Item Type
        5: 28,   # Title
        6: 55,   # Internal Path
        7: 55,   # Target
        8: 24,   # Impact area
        9: 90,   # Suggestion
        10: 18,  # Incidents Repeats
        11: 16,  # Review type
        12: 18,  # Developer
        13: 20,  # Iteration Path
    }

    _write_sheet(
        ws_details,
        DETAIL_HEADERS,
        detail_rows,
        col_widths=detail_widths,
        wrap_cols=[2, 3, 5],
    )

    _write_sheet(
        ws_azure,
        AZURE_HEADERS,
        azure_rows,
        col_widths=azure_widths,
        wrap_cols=[6, 7, 9],
    )

    wb.save(out_path)